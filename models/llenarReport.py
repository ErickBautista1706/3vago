from openpyxl import Workbook
from models.database import Database
from sqlalchemy import text

class LlenarReporte:
    def __init__(self, ruta_reporte):
        self.ruta_reporte = ruta_reporte
        self.libro = Workbook()
        self.hoja_usuarios = self.libro.active
        self.hoja_usuarios.title = "Hoja1"
        self.hoja_zonas = self.libro.create_sheet(title="Hoja2")

    def llenar_celda(self, hoja, celda, valor):
        hoja[celda] = valor

    def llenar_hoja_usuarios(self, registros):
        # Encabezados
        encabezados = ["ID", "Nombre", "Apellido Paterno", "Apellido Materno", "Alias", "Email", "Tipo Usuario"]
        for col_num, encabezado in enumerate(encabezados, 1):
            self.hoja_usuarios.cell(row=1, column=col_num, value=encabezado)

        # Datos
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                self.hoja_usuarios.cell(row=row_num, column=col_num, value=valor)

    def llenar_hoja_zonas(self, registros):
        # Encabezados
        encabezados = ["ID", "Nombre", "Ubicación", "Activo", "ID Usuario", "Fecha de Actualización"]
        for col_num, encabezado in enumerate(encabezados, 1):
            self.hoja_zonas.cell(row=1, column=col_num, value=encabezado)

        # Datos
        for row_num, registro in enumerate(registros, 2):
            for col_num, valor in enumerate(registro, 1):
                self.hoja_zonas.cell(row=row_num, column=col_num, value=valor)

    def guardar_reporte(self):
        self.libro.save(self.ruta_reporte)

    def tabla_usuarios():
            try:
                db = Database()
                conn = db.engine.connect()

                # Consulta a la tabla de usuarios
                query = text("""
                    SELECT u.id_usr, 
                        u.nombre,
                        u.apellidoP,
                        u.apellidoM,
                        u.alias,
                        u.email,
                        t.tipo_usr
                    FROM usuarios u
                    JOIN tipo_usuario t ON u.id_tuser = t.id_tpurs
                """)

                result = conn.execute(query)
                usuarios = result.fetchall()

                conn.close()

                return usuarios

            except Exception as e:
                print(f"Error al obtener la tabla de usuarios: {e}")
                return []

    def tabla_zonas():
            try:
                db = Database()
                conn = db.engine.connect()

                # Consulta a la tabla de zonas con JOIN en la tabla de usuarios
                query = text("""
                    SELECT z.id_zn,
                        z.nombre_zn,
                        z.ubicacion_zn,
                        z.activo_zn,
                        u.nombre || ' ' || u.apellidoP || ' ' || u.apellidoM AS nombre_usuario,
                        z.uptade_zn
                    FROM zonas z
                    JOIN usuarios u ON z.id_usr = u.id_usr
                """)

                result = conn.execute(query)
                zonas = result.fetchall()

                conn.close()

                return zonas

            except Exception as e:
                print(f"Error al obtener la tabla de zonas: {e}")
                return []